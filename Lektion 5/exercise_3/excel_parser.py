import openpyxl
from openpyxl.utils import get_column_letter
import csv
from openpyxl.chart import BarChart, Reference
import datetime

SALES_SHEET_NAME = "Sales"
REGION_SHEET_NAME = "Sales per region"

# Skapa en arbetsbok och välj aktivt kalkylblad
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "Sales"

# Skapa Sales per region sheet
sales_sheet = workbook[SALES_SHEET_NAME]
region_sheet = workbook.create_sheet(title=REGION_SHEET_NAME)


# Öppna CSV-filen och läs datan
with open("sales_data.csv") as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=",")
    data = list(csv_reader)

""" Total försäljning """

# Sätt rubrikerna på första raden
headers = data[0]
for column_index, header in enumerate(headers):
    worksheet.cell(row=1, column=column_index + 1, value=header)

data = sorted(data[1:], key=lambda x: datetime.datetime.strptime(x[0], "%Y-%m-%d"))

# Definiera hur stor vår tabell ska vara
table_range = f"A1:F{len(data)}"

# Skapa tabellen
table = openpyxl.worksheet.table.Table(displayName="SalesTable", ref=table_range)

# Sätt styles för tabellen
table.tableStyleInfo = openpyxl.worksheet.table.TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=True,
    showLastColumn=True,
    showRowStripes=True,
    showColumnStripes=False,
)

# Lägg till tabellen i vårt sheet
worksheet.add_table(table)

# Skriv in resten av datan i vårt table
for row_index, row_data in enumerate(data, start=1):
    for column_index, cell_data in enumerate(row_data):
        worksheet.cell(row=row_index + 1, column=column_index + 1, value=cell_data)

# Skapa en ny kolumn med total försäljning
worksheet.cell(row=1, column=6, value="Total försäljning")
for row_index in range(2, worksheet.max_row + 1):
    count = int(worksheet.cell(row=row_index, column=4).value)
    price = float(worksheet.cell(row=row_index, column=5).value)
    total_amount = count * price
    worksheet.cell(row=row_index, column=6, value=total_amount).number_format = "#,##0"

# Ändra stil på rubrikerna
for cell in worksheet[1]:
    cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=14)


""" PER REGION """

# Spara försäljning per region i en dict
region_sales = {}
for row_index in range(2, worksheet.max_row + 1):
    region = worksheet.cell(row=row_index, column=3).value
    total_sales = float(worksheet.cell(row=row_index, column=6).value)
    region_sales[region] = round(region_sales.get(region, 0) + total_sales, 2)

# Byt ark till region-arket
worksheet = workbook[region_sheet.title]

# Skapa rubriker
worksheet.cell(row=1, column=1, value="Region")
worksheet.cell(row=1, column=2, value="Sales")

# Formatera rubrikerna i feststil
for cell in worksheet[1]:
    cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF', size=14)

# Definiera hur stort vårt ark skall vara
table_range = f"A1:B{len(region_sales) + 1}"

# Skapa vår tabell
table = openpyxl.worksheet.table.Table(displayName="Region_Sales", ref=table_range)

# Ställ in tabellstilar
style = openpyxl.worksheet.table.TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
table.tableStyleInfo = style

# Lägg till tabellen till arket
worksheet.add_table(table)

# Skriv in datan för försäljning per region
for index, (name, sales) in enumerate(sorted(region_sales.items()), start=2):
    worksheet.cell(row=index, column=1, value=name)
    worksheet.cell(row=index, column=2, value=sales).number_format = "#,##0"

""" DIAGRAM """

# Skapa en datareferens till försäljningsdata
data = Reference(worksheet, min_col=2, min_row=1, max_col=2, max_row=worksheet.max_row)

# Skapa en kategorireferens till regioner
categories = Reference(worksheet, min_col=1, min_row=1, max_row=worksheet.max_row)

# Skapa ett stapeldiagram
chart = BarChart()
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
chart.title = "Försäljning per region"
chart.x_axis.title = "Region"
chart.y_axis.title = "Försäljning (kr)"

# Lägg till diagrammet i arbetsboken
worksheet.add_chart(chart, "D2")

# Spara arbetsboken
workbook.save("output.xlsx")
