import openpyxl
import csv
import datetime
import os

dir_path = os.path.dirname(__file__)
csv_path = os.path.join(dir_path, 'sales_data.csv')
output_path = os.path.join(os.path.dirname(__file__), 'output.xlsx')

SALES_SHEET_NAME = 'Sales_sheet'

# Skapa en arbetsbok och välj aktivt kalkylblad
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = SALES_SHEET_NAME

# Skapa Sales sheet
sales_sheet = workbook[SALES_SHEET_NAME]

# Öppna CSV-filen och läs datan
with open(csv_path) as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=',')
    data = list(csv_reader)

# Sätt rubrikerna på första raden
headers = data[0]
for column_index, header in enumerate(headers):
    worksheet.cell(row=1, column=column_index+1, value=str(header))

data = sorted(data[1:], key=lambda x: datetime.datetime.strptime(x[0], "%Y-%m-%d"))

# Skriv data till Excel resterande av dokumentet
for row_index, row_data in enumerate(data[1:], start=2):
    for column_index, cell_data in enumerate(row_data):
        worksheet.cell(row=row_index, column=column_index+1, value=cell_data)

# Lägg till kolumn med total försäljning
worksheet.cell(row=1, column=6, value='Total försäljning')
for row_index in range(2, worksheet.max_row + 1):
    count = int(worksheet.cell(row=row_index, column=4).value)
    price = float(worksheet.cell(row=row_index, column=5).value)
    total_amount = count * price
    worksheet.cell(row=row_index, column=6, value=total_amount).number_format = "#,##0"

# Formatera rubrikerna i feststil
for cell in worksheet[1]:
    cell.font = openpyxl.styles.Font(bold=True, size=14)

workbook.save(output_path)
