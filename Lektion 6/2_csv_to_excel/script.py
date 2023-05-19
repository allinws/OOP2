import openpyxl
import csv
import datetime
import os

dir_path = os.path.dirname(__file__)
csv_path = os.path.join(dir_path, 'sales_data.csv')
output_path = os.path.join(os.path.dirname(__file__), 'output.xlsx')

SALES_SHEET_NAME = 'Sales_sheet'

# Skapa en arbetsbok och välj aktivt kalkylblad
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = SALES_SHEET_NAME

# Skapa Sales sheet
sales_sheet = workbook['Sales_sheet']

# Öppna CSV-filen och läs datan
with open(csv_path) as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=',')
    data = list(csv_reader)

# Sätt rubrikerna på första raden
headers = data[0]

for column_index, header in enumerate(headers):
    worksheet.cell(row=1, column=column_index+1, value=header)

data = sorted(data[1:], key=lambda x: datetime.datetime.strptime(x[0], "%Y-%m-%d"))

# En lambda funktion är en funktion vi skriver på en rad
# nedan är en funktion som vi är vana vid att se den, dessa
# fungerar på samma sätt

# def sort_list(x):
#     datetime_x = datetime.datetime.strptime(x,  "%Y-%m-%d")
#     # sorterat efter datetime_x


# Skriv data till Excel resterande av dokumentet
for row_index, row_data in enumerate(data[1:], start=2):
    for column_index, cell_data in enumerate(row_data):
        worksheet.cell(row=row_index, column=column_index+1, value=cell_data)

workbook.save(output_path)
